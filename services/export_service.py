from database.db_manager import DatabaseManager
from services.student_service import StudentService
from services.class_service import ClassService
from services.subject_service import SubjectService
from services.grade_service import GradeService
from services.report_card_service import ReportCardService
from services.stats_service import StatsService

from openpyxl import Workbook
from collection import defaultdict

class ExportService:
    def __init__(self):
        self.db = DatabaseManager()
        self.students = StudentService()
        self.classes = ClassService()
        self.subjects = SubjectService()
        self.grades = GradeService()
        self.report = ReportCardService()
        self.stats = StatsService()

    def export_all_data(self, filename: str = "school_data.xlsx"):
        
        wb = Workbook()

        #students sheet
        ws_students = wb.active
        ws_students.title = "Students"
        students = self.students.get_all_students()
        ws_students.append(["ID", "Student ID", "Last Name", "First Name", "Gender", "Birth Date", "Class ID", "Parent Phone", "Registration Date", "Active"])
        for s in students:
            ws_students.append([
                s.id, s.student_id, s.last_name, s.first_name, s.gender, 
                s.birth_date, s.class_id, s.parent_phone, s.registration_date, s.is_active
            ]) 

        #classes sheet
        ws_classes = wb.create_sheet(title="Classes")
        classes = self.classes.get_all_classes()
        ws_classes.append(["ID", "Class Name", "Class Level", "School Year", "Maximum Students", "Creation Date"])
        for c in classes:
            ws_classes.append([
                c.id, c.class_name, c.class_level, c.school_year, 
                c.maximum_students, c.creation_date
            ])

        #subjects sheet
        ws_subjects = wb.create_sheet(title="Subjects")
        subjects = self.subjects.get_all_subjects()
        ws_subjects.append(["ID", "Subject Name", "Subject Weight", "Class ID"])
        for sub in subjects:
            ws_subjects.append([
                sub.id, sub.subject_name, sub.subject_weight, sub.class_id
            ])

        #grades sheet
        ws_grades = wb.create_sheet(title="Grades")
        grades = self.grades.get_all_grades()
        ws_grades.append(["ID", "Student ID", "Subject ID", "Score", "Type", "Grade Date", "Comment"])
        for g in grades:
            ws_grades.append([
                g.id, g.student_id, g.subject_id, g.score, g.evaluation_type, g.evaluation_date, g.comment
            ])
        
        wb.save(filename)
        return filename
    

    def export_report_card(self, student_id: str, filename: str = "report_card.xlsx"):

        report_card = self.report.generate_student_report(student_id)
        if not report_card:
            return None

        student = report_card["student"]
        grades = report_card["grades"]
        average = report_card["average"]
        class_average = report_card["class_average"]
        rank = report_card["rank"]
        total_students = report_card["total_students"]

        wb = Workbook()
        wb.title = f"Report Card - {student.first_name} {student.last_name}"
        ws = wb.active
        ws.append(["Report Card"])
        ws.append([])
        ws.append(["Student LastName", f"{student.last_name}"])
        ws.append(["Student FirstName", f"{student.first_name}"])
        ws.append(["Student ID", f"{student.student_id}"])
        ws.append(["Class ID", f"{student.class_id}"])
        ws.append([])

        ws.append(["Overall Average", f"{average:.2f}"])
        ws.append(["Class Average", f"{class_average:.2f}"])
        ws.append(["Rank", f"{rank}/{total_students}"])

        subject_scores = defaultdict(list)

        for g in grades:
            subject_scores[g.subject_id].append(g.score)

        ws.append([
            "Subject", "Weight", "Average Score", "Class Average", "Decision", "Honors"
        ])

        for subject_id, scores in subject_scores.items():
            
            subject = self.subjects.get_subject_by_id(subject_id)
            subject_avg = self.stats.get_subject_average(subject_id)
            student_subject_avg = sum(scores) / len(scores)
            decision = "Pass" if student_subject_avg >= 10 else "Fail"

            if student_subject_avg >= 18:
                honors = "A+"
            elif student_subject_avg >= 16:
                honors = "A"
            elif student_subject_avg >= 14:
                honors = "B"
            elif student_subject_avg >= 12:
                honors = "C"
            elif student_subject_avg >= 10:
                honors = "D"
            else:
                honors = "F"

            ws.append([
                subject.subject_name, subject.subject_weight, f"{student_subject_avg:.2f}", 
                f"{subject_avg:.2f}", decision, honors
            ])

            passed_subjects = 0
            failed_subjects = 0

            for score in subject_scores.values():
                avg_score = sum(score) / len(score)
                if avg_score >= 10:
                    passed_subjects += 1
                else:
                    failed_subjects += 1
        ws.append([])
        ws.append(["Rank", f"{rank}/{total_students}"])
        ws.append(["Overall Average", f"{average:.2f}"])
        ws.append(["Class Average", f"{class_average:.2f}"])
        ws.append(["Passed Subjects", f"{passed_subjects}"])
        ws.append(["Failed Subjects", f"{failed_subjects}"])
            