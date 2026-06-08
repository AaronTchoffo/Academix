from database.db_manager import DatabaseManager
from services.student_service import StudentService
from services.class_service import ClassService
from services.subject_service import SubjectService
from services.grade_service import GradeService

from openpyxl import load_workbook

class ImportService:
    def __init__(self):
        self.db = DatabaseManager()
        self.students = StudentService()
        self.classes = ClassService()
        self.subjects = SubjectService()
        self.grades = GradeService()

    def import_all_data(self, filename: str):
        wb = load_workbook(filename)

        #students sheet
        ws_students = wb["Students"]
        for row in ws_students.iter_rows(min_row=2, values_only=True):
            self.students.create_student(
                student_id=row[1], last_name=row[2], first_name=row[3],
                  gender=row[4], birth_date=row[5], class_id=row[6], parent_phone=row[7], registration_date=row[8], is_active=row[9])
            
        #classes sheet
        ws_classes = wb["Classes"]

        for row in ws_classes.iter_rows(min_row=2, values_only=True):
            self.classes.create_class(
                class_name=row[1], class_level=row[2], school_year=row[3], maximum_students=row[4], creation_date=row[5]
            )

        #subjects sheet
        ws_subjects = wb["Subjects"]
        for row in ws_subjects.iter_rows(min_row=2, values_only=True):
            self.subjects.create_subject(
                subject_name=row[1], subject_weight=row[2], class_id=row[3]
            )

        #grades sheet
        ws_grades = wb["Grades"]
        for row in ws_grades.iter_rows(min_row=2, values_only=True):
            self.grades.create_grade(
                student_id=row[1], subject_id=row[2], score=row[3], evaluation_type=row[4], evaluation_date=row[5], comment=row[6]
            )